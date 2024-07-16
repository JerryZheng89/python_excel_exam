from openpyxl import load_workbook

wb = load_workbook('./pyxl_write_test.xlsx')

ws = wb['Sheet1']

ws['A1'] = 'New Value'

wb.save('./pyxl_write_test.xlsx')
