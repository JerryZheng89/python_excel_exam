import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

# 填充一些示例数据
for row in range(1, 11):
    for col in range(1, 6):
        ws.cell(row=row, column=col, value=row*col)

# 创建一个条件格式化规则
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# 应用条件格式，如果单元格值大于20，则填充红色
rule = CellIsRule(operator='greaterThan', formula=['20'], fill=red_fill)
ws.conditional_formatting.add('A1:E10', rule)

# 保存工作簿
wb.save('conditional_formatting_example.xlsx')
